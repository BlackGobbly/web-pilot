#!/usr/bin/env python3
"""
Config-driven web scraping engine
==================================
Reads Excel config -> calls Node.js Playwright worker -> aggregates to Excel.

Usage:
  python scraper.py                              # Auto-find config.xlsx
  python scraper.py --excel websites.xlsx        # Specify config file
  python scraper.py --excel list.xlsx --pages 3  # Override pages per site
  python scraper.py --excel list.xlsx --output results.xlsx

Excel config columns (auto-detected from header row; English or Chinese headers accepted):
  A: URL          - Target page URL
  B: Source       - Site name / label
  C: Engine       - playwright(default; reserved for future use)
  D: Container    - List container CSS selector, e.g. #news-list
  E: Item         - Item row CSS selector, e.g. li.date
  F: Title        - Title CSS selector, e.g. a[href]
  G: Date         - Date CSS selector, e.g. .date-info
  H: Link         - Link CSS selector, e.g. a@href
  I: Date Format  - Y-m-d(default) / Y/m/d
  J: Pagination   - Next-page button CSS selector
  K: Pages        - Pages to scrape (default 1; 0 = all)
  L: Mode         - list(default) / table
  M: Preclick     - Pipe-separated tree node labels to expand
"""
import sys
import os
import json
import subprocess
import argparse
from collections import Counter

# ── Excel I/O ──
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)


def parse_args():
    parser = argparse.ArgumentParser(description='Config-driven web scraping engine')
    parser.add_argument('--excel', default='', help='Path to Excel config file')
    parser.add_argument('--output', default='', help='Output Excel file path')
    parser.add_argument('--pages', type=int, default=0, help='Pages per site (overrides config)')
    parser.add_argument('--debug', action='store_true', help='Debug mode: print parsed configs')
    return parser.parse_args()


def find_config(path=''):
    """Locate config Excel file. Checks specified path first, then common filenames."""
    if path:
        return path if os.path.exists(path) else ''
    for name in ['config.xlsx', 'websites.xlsx', 'urls.xlsx', 'sites.xlsx']:
        if os.path.exists(name):
            return name
    # Also check parent directory
    for name in ['config.xlsx', 'websites.xlsx']:
        p = os.path.join('..', name)
        if os.path.exists(p):
            return p
    return ''


def read_config(filepath):
    """Read Excel config and return a list of config dictionaries."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Column name mapping (English preferred, Chinese as fallback)
    KEYWORDS = {
        'url': ['url', '网址', '链接', '地址', 'address'],
        'source': ['source', '来源', '名称', '网站名', '站点'],
        'engine': ['engine', '引擎', 'method', '方式'],
        'container': ['container', 'list_container', '容器', '容器选择器', 'listcontainer'],
        'item': ['item', 'item_selector', '条目', '条目选择器', '行'],
        'title': ['title', 'title_selector', '标题', '标题选择器'],
        'date': ['date', 'date_selector', '日期', '时间', '日期选择器'],
        'link': ['link', 'link_selector', 'href', 'url_sel', '链接列'],
        'date_fmt': ['date_format', 'date_fmt', '日期格式', '格式'],
        'pagination': ['pagination', 'pagenav', '分页', '翻页', '下一页'],
        'pages': ['pages', 'max_pages', '页数', '抓取页数'],
        'mode': ['mode', 'type', '提取模式', '提取方式', '模式'],
        'preclick': ['preclick', 'treeclick', '预点击', '树点击', '展开'],
    }

    # Scan first 5 rows for headers
    col_map = {}
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for idx, val in enumerate(row):
            if val is None:
                continue
            val = str(val).strip().lower()
            for key, keywords in KEYWORDS.items():
                if key in col_map:
                    continue
                if val in [k.lower() for k in keywords]:
                    col_map[key] = idx
        if 'url' in col_map:
            break

    if 'url' not in col_map:
        print(f"[error] No URL column found! Headers: {[c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]}")
        return []

    configs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[col_map['url']]
        if not url:
            continue
        url = str(url).strip()
        if not url.startswith('http'):
            continue

        def get(key, default=''):
            if key not in col_map:
                return default
            val = row[col_map[key]]
            return str(val).strip() if val is not None else default

        cfg = {
            'url': url,
            'source': get('source', url),
            'engine': get('engine', 'playwright'),
            'container': get('container'),
            'item': get('item'),
            'title': get('title'),
            'date': get('date'),
            'link': get('link'),
            'date_fmt': get('date_fmt'),
            'pagination': get('pagination'),
            'pages': int(get('pages', '1')) if get('pages', '').isdigit() else 1,
            'mode': get('mode', 'list'),
            'preclick': get('preclick'),
        }
        configs.append(cfg)

    wb.close()
    return configs


def write_excel(items, output_path):
    """Write items to a formatted Excel file (Date | Source | Title | Link)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    hfont = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['Date', 'Source', 'Title', 'Link']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, ha, border

    dfont = Font(name='Calibri', size=10)
    lfont = Font(name='Calibri', size=10, color='0563C1', underline='single')
    da = Alignment(vertical='center', wrap_text=True)
    dla = Alignment(horizontal='center', vertical='center')
    lfill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

    for ri, item in enumerate(items, 2):
        c1 = ws.cell(row=ri, column=1, value=item.get('date', ''))
        c1.font, c1.alignment, c1.border = dfont, dla, border

        c2 = ws.cell(row=ri, column=2, value=item.get('source', ''))
        c2.font, c2.alignment, c2.border = dfont, da, border

        c3 = ws.cell(row=ri, column=3, value=item.get('title', ''))
        c3.font, c3.alignment, c3.border = dfont, da, border

        c4 = ws.cell(row=ri, column=4, value=item.get('link', ''))
        c4.font, c4.alignment, c4.border = lfont, da, border
        if item.get('link'):
            c4.hyperlink = item['link']

        if ri % 2 == 0:
            for c in range(1, 5):
                ws.cell(row=ri, column=c).fill = lfill

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 60
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    return output_path


def main():
    args = parse_args()

    # Find config
    config_path = find_config(args.excel)
    if not config_path:
        print("[error] No config file found!")
        print("   Specify: python scraper.py --excel your_file.xlsx")
        print("   Or place config.xlsx / websites.xlsx in the current directory")
        return 1

    print(f"[config] {os.path.abspath(config_path)}")
    configs = read_config(config_path)
    if not configs:
        print("[error] No valid URL configs found")
        return 1

    print(f"[data] {len(configs)} site(s) configured")

    # Apply --pages override
    if args.pages > 0:
        for cfg in configs:
            cfg['pages'] = args.pages

    if args.debug:
        for cfg in configs:
            print(f"  {cfg['source']}: {cfg['url']}")
            print(f"    container={cfg['container']} item={cfg['item']}")
            print(f"    title={cfg['title']} date={cfg['date']} link={cfg['link']}")

    # Launch Node.js worker
    workdir = os.path.dirname(os.path.abspath(__file__))
    worker_path = os.path.join(workdir, 'scraper_worker.js')

    if not os.path.exists(worker_path):
        print(f"[error] Worker not found: {worker_path}")
        return 1

    print(f"[start] Launching scraping engine...")
    proc = subprocess.Popen(
        ['node', worker_path],
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=workdir,
    )

    input_bytes = json.dumps(configs, ensure_ascii=False).encode('utf-8')
    timeout = 1800  # 30 min
    sys.stderr.write(f'  Timeout: {timeout}s\n')
    stdout_bytes, stderr_bytes = proc.communicate(input_bytes, timeout=timeout)

    # Show worker logs
    if stderr_bytes:
        try:
            sys.stderr.buffer.write(stderr_bytes)
        except:
            pass

    if proc.returncode != 0:
        print(f"[error] Worker exited abnormally (code={proc.returncode})")
        return 1

    try:
        all_items = json.loads(stdout_bytes.decode('utf-8'))
    except json.JSONDecodeError as e:
        print(f"[error] Failed to parse worker output: {e}")
        preview = stdout_bytes.decode('utf-8', errors='replace')[:500]
        print(f"   Raw output: {preview}")
        return 1

    if not all_items:
        print("[warning] No items scraped")
        return 0

    # Sort by date descending
    all_items.sort(key=lambda x: x.get('date', ''), reverse=True)

    # Summary
    print(f"\n{'='*50}")
    print(f"[data] Summary")
    print(f"{'='*50}")
    print(f"   Total items: {len(all_items)}")
    for src, cnt in Counter(i.get('source', 'Unknown') for i in all_items).most_common():
        print(f"   {src}: {cnt} items")
    print(f"   Date range: {all_items[-1].get('date','?')} ~ {all_items[0].get('date','?')}")

    # Output
    default_out = args.output or config_path.replace('.xlsx', '_output.xlsx')
    if default_out == config_path:
        default_out = config_path.rsplit('.', 1)[0] + '_output.xlsx'

    out_path = write_excel(all_items, default_out)
    print(f"\n[saved] {os.path.abspath(out_path)}")
    return 0


if __name__ == '__main__':
    sys.exit(main())
